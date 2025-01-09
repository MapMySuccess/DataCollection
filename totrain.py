import requests
import openpyxl
import time
import json
import math

# Your Google API Key
GOOGLE_MAPS_API_KEY = 'AIzaSyDYfzrpjY2RB7h9Bq8Q3sQ-bKjNDTFFKdE'

# Load data from data.json
with open('data.json', 'r') as json_file:
    place_data = json.load(json_file)

# Function to get nearby establishments within 1 km, including paginated results
def get_nearby_establishments(lat, lng, radius=500):
    places = []
    url = f"https://maps.googleapis.com/maps/api/place/nearbysearch/json?location={lat},{lng}&radius={radius}&key={GOOGLE_MAPS_API_KEY}"

    while url:
        response = requests.get(url)
        if response.status_code == 200:
            result = response.json()
            places.extend(result.get('results', []))  # Add new places to the list

            # Check if there's a next_page_token
            next_page_token = result.get('next_page_token')
            if next_page_token:
                # Update the URL for the next request with the next_page_token
                url = f"https://maps.googleapis.com/maps/api/place/nearbysearch/json?pagetoken={next_page_token}&key={GOOGLE_MAPS_API_KEY}"
                time.sleep(2)  # Required delay before requesting the next page
            else:
                # No more pages to fetch
                url = None
        else:
            # If the API request fails, stop further requests
            break

    return places

# Function to map Google's place types to a broader category found in the data.json file
def map_to_broader_category(place_types):
    # Normalize the keys in the JSON file by making them lowercase and replacing spaces with underscores
    normalized_json_keys = {key.lower().replace(" ", "_"): value for key, value in place_data.items()}

    # We'll check if any place types map directly to our normalized JSON categories
    for place_type in place_types:
        broader_category = normalized_json_keys.get(place_type, None)  # Use normalized key lookup
        if broader_category is not None:
            return place_type  # Return the first matched broader category
    return None

# Function to get numeric value for a place category from the normalized JSON data
def get_numeric_value_for_place(place_category):
    # Look up the place category in the normalized JSON data
    normalized_json_keys = {key.lower().replace(" ", "_"): value for key, value in place_data.items()}
    return normalized_json_keys.get(place_category, None)

# Function to calculate average traffic using Google Maps Distance Matrix API
def calculate_average_traffic(lat, lng, radius=500):
    # Define destinations for traffic checks (some random points within 1km)
    destinations = [
        f"{lat + 0.005},{lng}",
        f"{lat - 0.005},{lng}",
        f"{lat},{lng + 0.005}",
        f"{lat},{lng - 0.005}"
    ]
    
    url = f"https://maps.googleapis.com/maps/api/distancematrix/json?origins={lat},{lng}&destinations={'|'.join(destinations)}&departure_time=now&key={GOOGLE_MAPS_API_KEY}"
    
    # Make the API request
    response = requests.get(url)
    
    if response.status_code == 200:
        data = response.json()
        
        travel_times = []
        
        # Iterate through the response rows and elements
        for row in data.get('rows', []):
            for element in row.get('elements', []):
                # Try to get 'duration_in_traffic', fallback to 'duration' if not available
                if 'duration_in_traffic' in element:
                    travel_times.append(element['duration_in_traffic']['value'])  # Get traffic time in seconds
                elif 'duration' in element:
                    travel_times.append(element['duration']['value'])  # Fallback to normal travel time
        
        if travel_times:
            avg_traffic_time = sum(travel_times) / len(travel_times)  # Calculate average in seconds
            return avg_traffic_time / 60  # Convert to minutes
        else:
            return None
    else:
        return None

# Function to convert traffic time to traffic severity (1-5 scale)
def convert_traffic_time_to_severity(avg_traffic_time):
    if avg_traffic_time is None:
        return 0  # No traffic data
    if avg_traffic_time < 5:
        return 1  # Low traffic
    elif avg_traffic_time < 10:
        return 2  # Moderate traffic
    elif avg_traffic_time < 15:
        return 3  # Heavy traffic
    elif avg_traffic_time < 20:
        return 4  # Very heavy traffic
    else:
        return 5  # Severe traffic

# Function to find the distance to the nearest main road using Google Maps Roads API
def find_distance_to_nearest_main_road(lat, lng):
    url = f"https://roads.googleapis.com/v1/nearestRoads?points={lat},{lng}&key={GOOGLE_MAPS_API_KEY}"
    response = requests.get(url)
    
    if response.status_code == 200:
        result = response.json()
        if 'snappedPoints' in result:
            nearest_road_location = result['snappedPoints'][0]['location']
            nearest_lat = nearest_road_location['latitude']
            nearest_lng = nearest_road_location['longitude']
            
            # Use Haversine formula to calculate distance to the nearest road
            distance = calculate_distance(lat, lng, nearest_lat, nearest_lng)
            return distance
        else:
            return None
    else:
        return None

# Function to calculate the distance between two latitude and longitude points (Haversine formula)
def calculate_distance(lat1, lon1, lat2, lon2):
    R = 6371  # Earth radius in kilometers
    d_lat = math.radians(lat2 - lat1)
    d_lon = math.radians(lon2 - lon1)
    a = math.sin(d_lat/2) * math.sin(d_lat/2) + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(d_lon/2) * math.sin(d_lon/2)
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    distance = R * c  # Distance in kilometers
    return distance * 1000  # Convert to meters

# Function to write the calculated data into the Excel file
def write_data_to_excel(file_path, row_num, avg_pdensity, traffic_rate, visibility, comp_presence, average_price_level):
    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Write the calculated values into the respective columns
    sheet[f'G{row_num}'] = avg_pdensity  # Assuming AVG_PDENSITY is in column G
    sheet[f'H{row_num}'] = traffic_rate  # Assuming TRAFFIC_RATE is in column H
    sheet[f'I{row_num}'] = visibility  # Assuming VISIBILITY is in column I
    sheet[f'J{row_num}'] = comp_presence  # Assuming COMPETITION_PRESENCE is in column J
    sheet[f'K{row_num}'] = average_price_level  # Assuming COMPETITION_PRESENCE is in column J

    # Save the workbook to preserve changes
    workbook.save(file_path)

# Function to process each row of the Excel file and extract the latitude and longitude
def process_excel_file(file_path):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Iterate over the rows in the Excel sheet
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):  # Skip header row and start from row 2
        print(row)
        name, address, rating, rating_total, latitude, longitude = row
        # Ensure latitude and longitude are present
        if latitude and longitude:
            print(f"Processing {name}: Lat {latitude}, Lng {longitude}")
            # Get the calculated metrics
            avg_pdensity, traffic_rate, visibility ,comp_presence, average_price_level= find_restaurant_details(latitude, longitude, name)

            # Write the calculated data to the Excel file
            write_data_to_excel(file_path, idx, avg_pdensity, traffic_rate, visibility, comp_presence, average_price_level)



# Function to get nearby restaurants within 500 meters
def get_nearby_restaurants(lat, lng, radius=5000):
    restaurants = []
    url = f"https://maps.googleapis.com/maps/api/place/nearbysearch/json?location={lat},{lng}&radius={radius}&type=restaurant&key={GOOGLE_MAPS_API_KEY}"
    
    while url:
        response = requests.get(url)
        if response.status_code == 200:
            result = response.json()
            places = result.get('results', [])
            restaurants.extend(places)  # Add new restaurants to the list

            # Check if there's a next_page_token
            next_page_token = result.get('next_page_token')
            if next_page_token:
                url = f"https://maps.googleapis.com/maps/api/place/nearbysearch/json?pagetoken={next_page_token}&key={GOOGLE_MAPS_API_KEY}"
                time.sleep(2)  # Required delay before requesting the next page
            else:
                # No more pages to fetch
                url = None
        else:
            # If the API request fails, stop further requests
            break

    return restaurants

# Function to filter same-type restaurants
def filter_same_type_restaurants(restaurants, restaurant_type):
    same_type_restaurants = []
    for restaurant in restaurants:
        types = restaurant.get('types', [])
        if restaurant_type in types:
            same_type_restaurants.append(restaurant)
    return same_type_restaurants


# Function to calculate competitor presence
def calculate_competitor_presence(total_restaurants, same_type_restaurants):
    # Total area of a 500m radius circle (fixed value)
    AREA_500M_RADIUS = math.pi * (5000 ** 2)

    # If no restaurants are found, return the lowest competition score
    if total_restaurants == 0:
        return 1  # No competition

    # Calculate the competitor ratio (same-type to total)
    competitor_ratio = same_type_restaurants / total_restaurants

    # Calculate restaurant density
    restaurant_density = total_restaurants / AREA_500M_RADIUS

    # Calculate competition score
    competition_score = (competitor_ratio * restaurant_density)

    # Normalize the score to a range from 1 (low competition) to 5 (high competition)
    return normalize_score(competition_score)

# Function to normalize the competition score to a 1-5 scale
def normalize_score(competition_score):
    # Example thresholds (can be adjusted based on data insights)
    if competition_score < 0.1:
        return 1  # Low competition
    elif competition_score < 0.3:
        return 2  # Slight competition
    elif competition_score < 0.5:
        return 3  # Medium competition
    elif competition_score < 0.7:
        return 4  # High competition
    else:
        return 5  # Very high competition

# Function to calculate competitor presence for a given location
def competitor_presence_for_location(lat, lng, restaurant_type):
    # Get nearby restaurants
    restaurants = get_nearby_restaurants(lat, lng)

    # Filter same-type restaurants
    same_type_restaurants = filter_same_type_restaurants(restaurants, restaurant_type)

    # Total number of restaurants
    total_restaurants = len(restaurants)

    # Same-type restaurants count
    same_type_restaurant_count = len(same_type_restaurants)

    # Calculate competitor presence
    competitor_presence = calculate_competitor_presence(total_restaurants, same_type_restaurant_count)

    return competitor_presence




# Function to find a restaurant's price range using the Google Places API
def get_place_id(place_name, api_key):
    """
    Use Google Places API to find the place_id for a given place name.
    """
    url = f"https://maps.googleapis.com/maps/api/place/findplacefromtext/json?input={place_name}&inputtype=textquery&key={api_key}"
    response = requests.get(url)
    
    if response.status_code == 200:
        data = response.json()
        if data.get('candidates'):
            return data['candidates'][0]['place_id']  # Return the first match place_id
        else:
            return None
    else:
        print(f"Error fetching place_id for {place_name}: {response.status_code}")
        return None

def get_place_details(place_id, api_key):
    """
    Use Google Places API to get the details of a place by place_id.
    """
    url = f"https://maps.googleapis.com/maps/api/place/details/json?place_id={place_id}&fields=price_level&key={api_key}"
    response = requests.get(url)
    
    if response.status_code == 200:
        data = response.json()
        if 'result' in data:
            # Return the price level, or None if not available
            return data['result'].get('price_level', 'N/A')
        else:
            return 'N/A'
    else:
        print(f"Error fetching place details for place_id {place_id}: {response.status_code}")
        return 'N/A'

# Function to read restaurant names from an Excel file
def read_place_names_from_excel(file_path):
    """
    Read the place names from the Excel file.
    """
    place_names = []
    
    # Load the workbook and the first sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    # Assuming the place names are in the first column (A)
    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):  # Skip the header
        place_name = row[0]  # Get the place name from column A
        if place_name:
            place_names.append(place_name)
    
    return place_names

# Function to update the price range in the Excel file
def update_excel_with_price_ranges(file_path, price_ranges):
    """
    Write the price range values into the Excel file.
    """
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    # Assuming price range should be written to column B (next to place names)
    for idx, price_range in enumerate(price_ranges, start=2):
        sheet[f'B{idx}'] = price_range  # Write price range to column B
    
    # Save the updated workbook
    workbook.save(file_path)

# Main function to process the data
def process_places_and_update_excel(file_path, api_key):
    """
    Main function that reads the Excel file, fetches price range, and updates the Excel.
    """
    # Read place names from the Excel file
    place_names = read_place_names_from_excel(file_path)
    
    # List to store price ranges
    price_ranges = []
    
    # Iterate over each place name and fetch its price range
    for place_name in place_names:
        print(f"Fetching price range for: {place_name}")
        
        # Get place_id using the place name
        place_id = get_place_id(place_name, api_key)
        
        if place_id:
            # Get the place details (price range) using the place_id
            price_range = get_place_details(place_id, api_key)
        else:
            price_range = 'Not Found'
        
        # Append the price range to the list
        price_ranges.append(price_range)
    
    # Update the Excel file with the retrieved price ranges
    update_excel_with_price_ranges(file_path, price_ranges)
    print(f"Updated price ranges in {file_path}.")








# Main function to handle user input and calculate required metrics
def find_restaurant_details(lat, lng, restaurant_type):
    # Get nearby establishments within 1 km
    nearby_establishments = get_nearby_establishments(lat, lng)

    # Initialize a list to store unique place categories
    places = []
    
    # Initialize a list to store numeric values corresponding to place categories
    numbers = []

    # Initialize a list to store price levels of restaurants
    price_levels = []

    # Iterate through the nearby establishments and check their types against the JSON file
    for place in nearby_establishments:
        place_types = place.get('types', [])  # List of types for the place
        
        # Try to map the place types to a broader category (only once per place)
        broader_category = map_to_broader_category(place_types)

        if broader_category:
            # Add unique broader category to the places list
            if broader_category not in places:
                places.append(broader_category)
            
            # Get the numeric value for this broader category from the JSON file
            numeric_value = get_numeric_value_for_place(broader_category)
            
            # Only add one numeric value per place
            if numeric_value is not None and broader_category not in numbers:
                numbers.append(int(numeric_value))  # Ensure numeric value is an integer

            # Fetch the price level for restaurant-type places
            if 'restaurant' in place_types:
                place_id = place.get('place_id')
                if place_id:
                    price_level = get_place_details(place_id, GOOGLE_MAPS_API_KEY)
                    
                    # Ensure price_level is numeric (convert if needed) or append 0 if it's invalid
                    try:
                        if price_level is not None and price_level != 'N/A':
                            # Ensure price_level is an integer
                            price_levels.append(int(price_level))
                        else:
                            # Append a default value of 0 if price_level is 'None' or 'N/A'
                            price_levels.append(0)
                    except ValueError:
                        # If price_level is not a valid integer, append 0
                        price_levels.append(0)


    # Calculate the average population density based on the numbers list
    if numbers:
        Avg_population_density = sum(numbers) / len(numbers)
    else:
        Avg_population_density = 0

    # Get average traffic in the area
    avg_traffic = calculate_average_traffic(lat, lng)

    # Convert traffic time to a severity rating (1-5)
    traffic_severity = convert_traffic_time_to_severity(avg_traffic)

    # Find distance to nearest main road
    distance_to_main_road = find_distance_to_nearest_main_road(lat, lng)

    # Calculate competitor presence for the given coordinates and restaurant type
    competitor_presence = competitor_presence_for_location(lat, lng, restaurant_type)


    if price_levels:
        average_price_level = sum(price_levels) / len(price_levels)
    else:
        average_price_level = None  # No price levels available

    
    # Return the calculated metrics
    return Avg_population_density, traffic_severity, distance_to_main_road, competitor_presence, average_price_level

# Provide the Excel file path
file_path = "all_nearby_places1.xlsx"
process_excel_file(file_path)
